from flask import Flask, render_template, request, send_file
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile, os, uuid
import matplotlib
matplotlib.use('Agg')  # Use non-GUI backend
import matplotlib.pyplot as plt

app = Flask(__name__)

@app.route('/')
def index():
    return render_template("index.html")

@app.route('/process', methods=['POST'])
def process():
    file = request.files['file']
    if not file or not file.filename.endswith('.xlsx'):
        return "Upload a valid Excel file.", 400

    input_temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    file.save(input_temp.name)
    output_temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    output_file = output_temp.name

    xls = pd.ExcelFile(input_temp.name)
    wb = Workbook()
    wb.remove(wb.active)
    combined_data = []

    for sheet in xls.sheet_names:
        df = xls.parse(sheet)
        if not {'DATE', 'WIDTH', 'OLD PRICE'}.issubset(df.columns):
            continue
        df['OLD PRICE'] = df['OLD PRICE'].astype(str).str.replace('₹', '').str.replace(',', '')
        df['OLD PRICE'] = pd.to_numeric(df['OLD PRICE'], errors='coerce')
        df.dropna(subset=['DATE', 'WIDTH', 'OLD PRICE'], inplace=True)
        combined_data.append(df[['DATE', 'WIDTH', 'OLD PRICE']])
        pivot = pd.pivot_table(df, index='WIDTH', columns='DATE', values='OLD PRICE', aggfunc='mean').reset_index()
        ws = wb.create_sheet(title=sheet[:31])
        for row in dataframe_to_rows(pivot, index=False, header=True):
            ws.append(row)

    if combined_data:
        total = pd.concat(combined_data)
        total_pivot = pd.pivot_table(total, index='WIDTH', columns='DATE', values='OLD PRICE', aggfunc='mean').reset_index()
        date_cols = total_pivot.columns[1:]
        sorted_dates = sorted(date_cols, key=lambda x: pd.to_datetime(x, errors='coerce'))

        if len(sorted_dates) >= 2:
            prev, curr = sorted_dates[-2], sorted_dates[-1]
            total_pivot['% CHANGE'] = ((total_pivot[curr] / total_pivot[prev]) - 1) * 100
            total_pivot['% CHANGE'] = total_pivot['% CHANGE'].round(2).astype(str) + '%'
        else:
            total_pivot['% CHANGE'] = 'NA'
            prev = curr = None

        ws_total = wb.create_sheet(title='TOTAL')
        for row in dataframe_to_rows(total_pivot, index=False, header=True):
            ws_total.append(row)

        if prev and curr:
            summary = pd.DataFrame({
                'WIDTH': total_pivot['WIDTH'],
                'PREVIOUS MONTH VALUE': total_pivot[prev],
                'PRESENT MONTH VALUE': total_pivot[curr],
                '% CHANGE': total_pivot['% CHANGE']
            })
            ws_sum = wb.create_sheet(title='PERCENTAGE CHANGE SUMMARY')
            for row in dataframe_to_rows(summary, index=False, header=True):
                ws_sum.append(row)

    wb.save(output_file)
    return send_file(output_file, as_attachment=True, download_name='Pivot_Report.xlsx')

@app.route('/visualize', methods=['POST'])
def visualize():
    file = request.files['file']
    if not file or not file.filename.endswith('.xlsx'):
        return "Invalid file", 400

    temp_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    file.save(temp_path)

    xls = pd.ExcelFile(temp_path)
    combined = []

    for sheet in xls.sheet_names:
        df = xls.parse(sheet)
        if {'DATE', 'WIDTH', 'OLD PRICE'}.issubset(df.columns):
            df['OLD PRICE'] = df['OLD PRICE'].astype(str).str.replace('₹', '').str.replace(',', '')
            df['OLD PRICE'] = pd.to_numeric(df['OLD PRICE'], errors='coerce')
            df = df.dropna(subset=['DATE', 'WIDTH', 'OLD PRICE'])
            combined.append(df[['DATE', 'WIDTH', 'OLD PRICE']])

    if not combined:
        return "No valid data found"

    total = pd.concat(combined)
    latest_date = total['DATE'].dropna().sort_values().unique()[-1]
    latest = total[total['DATE'] == latest_date]
    group = latest.groupby('WIDTH')['OLD PRICE'].mean().sort_index()

    filename = f"{uuid.uuid4().hex}.png"
    img_path = os.path.join("static", filename)

    plt.figure(figsize=(10, 6))
    group.plot(kind='bar', color='orange')
    plt.title(f"OLD PRICE by WIDTH - {latest_date}")
    plt.ylabel("Price")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig(img_path)
    plt.close()

    return render_template("visualize.html", image_path=img_path, date=latest_date)

@app.route('/table', methods=['POST'])
def table():
    file = request.files['file']
    if not file or not file.filename.endswith('.xlsx'):
        return "Invalid file", 400

    temp_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    file.save(temp_path)
    xls = pd.ExcelFile(temp_path)
    combined = []

    for sheet in xls.sheet_names:
        df = xls.parse(sheet)
        if {'DATE', 'WIDTH', 'OLD PRICE'}.issubset(df.columns):
            df['OLD PRICE'] = df['OLD PRICE'].astype(str).str.replace('₹', '').str.replace(',', '')
            df['OLD PRICE'] = pd.to_numeric(df['OLD PRICE'], errors='coerce')
            df = df.dropna(subset=['DATE', 'WIDTH', 'OLD PRICE'])
            combined.append(df[['DATE', 'WIDTH', 'OLD PRICE']])

    if not combined:
        return "No valid data to show."

    final = pd.concat(combined)
    last = final['DATE'].dropna().sort_values().unique()[-1]
    last_data = final[final['DATE'] == last]
    table_data = last_data.groupby('WIDTH')['OLD PRICE'].mean().reset_index().sort_values('WIDTH')

    return render_template("table.html", data=table_data.to_dict(orient='records'), date=last)

if __name__ == '__main__':
    app.run(debug=True)
